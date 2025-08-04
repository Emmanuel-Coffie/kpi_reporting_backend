# management/commands/populate_kpis.py
import os
from django.core.management.base import BaseCommand
from django.db import transaction
from datetime import date
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.styles import numbers
from ...models import Directorate, KPIYear, PredefinedKPI, MonthlyTarget, Initiative

class Command(BaseCommand):
    help = 'Populates the database with Directorates, KPIs, Initiatives and Targets from Excel data'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Path to the Excel file',
            default='Initiatives-KPIs mapping.xlsx'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        try:
            wb = load_workbook(filename=file_path, data_only=False)  # data_only=False to get cell formatting
            sheet = wb['Directorates_KPIs']  # Use the correct sheet name
            
            with transaction.atomic():
                # 1. Create current KPI Year
                current_year = date.today().year
                kpi_year, _ = KPIYear.objects.get_or_create(
                    year=current_year,
                    defaults={'is_current': True}
                )

                # 2. Get all rows from Excel (skip header row)
                rows = list(sheet.iter_rows(min_row=2))  # Skip header row

                # 3. Create all Directorates
                directorates = set(row[0].value for row in rows if row[0].value)  # Column A
                directorate_objs = {}
                for name in directorates:
                    if name:  # Skip empty names
                        obj, _ = Directorate.objects.get_or_create(name=name)
                        directorate_objs[name] = obj

                # 4. Create all KPIs and Initiatives
                kpi_cache = {}  # (directorate_name, kpi_name) -> KPI object
                initiative_cache = {}  # (directorate_name, initiative_name) -> Initiative object

                for row in rows:
                    directorate_name = row[0].value  # Column A
                    corporate_kpi_name = row[1].value  # Column B
                    initiative_name = row[2].value  # Column C
                    directorate_kpi_name = row[3].value  # Column D
                    
                    # Skip rows without proper KPI name
                    if not directorate_kpi_name or not isinstance(directorate_kpi_name, str):
                        continue
                    
                    # Get directorate object
                    directorate = directorate_objs.get(directorate_name)
                    if not directorate:
                        continue
                    
                    # Create or get KPI
                    is_corporate = (directorate_name == 'Corporate')
                    kpi_key = (directorate_name, directorate_kpi_name.strip())
                    
                    if kpi_key not in kpi_cache:
                        try:
                            unit = str(row[4].value) if row[4].value else ''  # Column E
                            kpi = PredefinedKPI.objects.create(
                                kpi_year=kpi_year,
                                directorate=directorate,
                                name=directorate_kpi_name.strip(),
                                unit_of_measurement=unit,
                                aggregation_type='AVG' if (row[5].value or '').upper() == 'A' else 'SUM',  # Column F
                                performance_logic='HIGHER' if (str(row[6].value or '').upper() != 'L') else 'LOWER',  # Column G
                                is_corporate=is_corporate
                            )
                            kpi_cache[kpi_key] = kpi
                        except Exception as e:
                            self.stdout.write(self.style.ERROR(f"Error creating KPI {directorate_kpi_name}: {str(e)}"))
                            continue
                    
                    # Create or get Initiative (for non-corporate rows)
                    if not is_corporate and initiative_name and isinstance(initiative_name, str):
                        initiative_key = (directorate_name, initiative_name.strip())
                        
                        if initiative_key not in initiative_cache:
                            try:
                                initiative = Initiative.objects.create(
                                    directorate=directorate,
                                    kpi_year=kpi_year,
                                    title=initiative_name.strip(),
                                    start_date=date(current_year, 1, 1),
                                    end_date=date(current_year, 12, 31)
                                )
                                initiative_cache[initiative_key] = initiative
                            except Exception as e:
                                self.stdout.write(self.style.ERROR(f"Error creating initiative {initiative_name}: {str(e)}"))
                                continue
                        
                        # Link KPI to initiative
                        initiative = initiative_cache[initiative_key]
                        kpi = kpi_cache[kpi_key]
                        
                        try:
                            if is_corporate:
                                initiative.corporate_kpis.add(kpi)
                            else:
                                initiative.linked_kpis.add(kpi)
                        except Exception as e:
                            self.stdout.write(self.style.ERROR(f"Error linking KPI to initiative: {str(e)}"))
                            continue
                    
                    # Create Monthly Targets for each KPI (columns J-S, months 1-12)
                    kpi = kpi_cache.get(kpi_key)
                    if not kpi:
                        continue
                    
                    # Check if unit is percentage to handle differently
                    is_percentage = kpi.unit_of_measurement == '%'
                    
                    for month_idx, cell in enumerate(row[9:21], start=1):  # Columns J-S
                        target_value = cell.value
                        if target_value in ['N/A', 'n/a', None, '']:
                            continue
                        
                        try:
                            # Check if cell is formatted as percentage
                            cell_is_percentage = cell.number_format and ('%' in cell.number_format or 
                                                                        cell.number_format in numbers.FORMAT_PERCENTAGE)
                            
                            # Handle different numeric formats
                            if isinstance(target_value, (int, float)):
                                if cell_is_percentage and is_percentage:
                                    # Convert percentage-formatted value to whole number (0.98 → 98)
                                    numeric_value = Decimal(str(target_value * 100))
                                else:
                                    numeric_value = Decimal(str(target_value))
                            elif isinstance(target_value, str):
                                if target_value.startswith('='):  # Skip formulas
                                    continue
                                # Remove any commas and percentage signs
                                clean_value = target_value.replace(',', '').replace('%', '')
                                numeric_value = Decimal(clean_value)
                            else:
                                continue
                            
                            # For percentages, keep as whole numbers (e.g., 98% → 98, not 0.98)
                            if is_percentage:
                                numeric_value = round(numeric_value, 2)
                            else:
                                # For non-percentages, round to 2 decimal places
                                numeric_value = round(numeric_value, 2)
                            
                            # Negative values are handled naturally by Decimal
                            MonthlyTarget.objects.get_or_create(
                                predefined_kpi=kpi,
                                year=current_year,
                                month=month_idx,
                                defaults={'target_value': numeric_value}
                            )
                        except Exception as e:
                            self.stdout.write(self.style.WARNING(
                                f"Skipping target for {directorate_kpi_name} month {month_idx}: {str(e)}"
                            ))
                            continue

            self.stdout.write(self.style.SUCCESS('Successfully populated database'))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error processing Excel file: {str(e)}"))
            raise e